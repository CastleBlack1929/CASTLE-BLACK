#!/usr/bin/env python3
"""
sync_excel.py — Sincroniza datos del Excel de Castle Black al dashboard.

Qué sincroniza:
  1. MOVIMIENTOS  → nuevas filas del Excel → data/movimientos.js
  2. COMPORTAMIENTO → patrimonio mensual por cliente → data/usuarios/*.js

Uso:
  python3 scripts/sync_excel.py
"""

import json
import re
import os
import warnings
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl. Instala con:  pip3 install openpyxl")
    raise

# ── Rutas ─────────────────────────────────────────────────────────────────────

EXCEL_PATH = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/"
    "ARCHIVO/REGISTRO CASTLE BLACK/2026/Castle-Black investments 2026.xlsx"
)
REPO_PATH      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MOVIMIENTOS_JS = os.path.join(REPO_PATH, "data", "movimientos.js")
USERS_JS       = os.path.join(REPO_PATH, "data", "users.js")

MESES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]

# ── Parseo JS ─────────────────────────────────────────────────────────────────

def _extract_between(text, open_char, close_char):
    start = text.index(open_char)
    depth = 0
    for i, ch in enumerate(text[start:], start):
        if ch == open_char:
            depth += 1
        elif ch == close_char:
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    raise ValueError(f"No se encontró '{close_char}' balanceado")

def parse_movimientos(path):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    return json.loads(_extract_between(text, "[", "]"))

def parse_userdata(path):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    return json.loads(_extract_between(text, "{", "}"))

def parse_users(path):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    return json.loads(_extract_between(text, "[", "]"))

# ── Escritura JS ──────────────────────────────────────────────────────────────

def _fmt_val(v):
    if isinstance(v, str):
        return f'"{v.replace(chr(92), chr(92)*2).replace(chr(34), chr(92)+chr(34))}"'
    if isinstance(v, float):
        return f"{v}" if v != int(v) else f"{v:.1f}"
    if isinstance(v, bool):
        return "true" if v else "false"
    if v is None:
        return "null"
    return str(v)

def write_movimientos(path, data):
    lines = ["const movimientosData = ["]
    for ei, entry in enumerate(data):
        lines.append("  {")
        items = list(entry.items())
        for fi, (k, v) in enumerate(items):
            comma = "," if fi < len(items) - 1 else ""
            lines.append(f'    "{k}": {_fmt_val(v)}{comma}')
        comma = "," if ei < len(data) - 1 else ""
        lines.append("  }" + comma)
    lines.append("];")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

def write_userdata(path, data):
    with open(path, "w", encoding="utf-8") as f:
        f.write("window.userData = ")
        f.write(json.dumps(data, ensure_ascii=False, indent=2))
        f.write(";\n")

# ── Helpers ───────────────────────────────────────────────────────────────────

def build_users_map(users_list):
    mapping = {}
    for u in users_list:
        m = re.search(r"/(\w+)\.js$", u.get("dataFile", ""))
        if m:
            mapping[m.group(1)] = {"username": u["username"], "dataFile": u["dataFile"]}
    return mapping

def fmt_fecha(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%y")
    return str(value).strip()

def fmt_cliente(n_raw):
    try:
        return str(int(n_raw)).zfill(3)
    except (TypeError, ValueError):
        return str(n_raw)

def safe_float(v):
    try:
        return round(float(v), 5)
    except (TypeError, ValueError):
        return 0.0

def n_key_of(n_raw):
    try:
        return str(int(n_raw))
    except (TypeError, ValueError):
        return str(n_raw).strip()

# ── Sync 1: MOVIMIENTOS ───────────────────────────────────────────────────────

def sync_movimientos(ws, mov_data, users_map):
    existing = {str(m.get("recibo", "")) for m in mov_data}
    nuevos = []

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
        recibo_raw = row[1]
        fecha_raw  = row[2]
        n_raw      = row[3]
        socio      = row[4]
        cedula     = row[5]
        cantidad   = row[6]
        tipo       = row[7]
        tasa_raw   = row[8]
        cambio_raw = row[9]

        if recibo_raw is None or socio is None or cantidad is None:
            continue
        recibo_str = str(int(recibo_raw)) if isinstance(recibo_raw, (int, float)) else str(recibo_raw).strip()
        if not recibo_str.isdigit():
            continue
        if recibo_str in existing:
            continue

        nk = n_key_of(n_raw)

        # Castle Black se genera automáticamente — no importar
        if nk == "A" or str(socio).strip().upper() == "CASTLE BLACK":
            continue

        user_info = users_map.get(nk, {})
        entry = {
            "username" : user_info.get("username", "MATRIX"),
            "cliente"  : fmt_cliente(n_raw),
            "recibo"   : recibo_str,
            "fecha"    : fmt_fecha(fecha_raw),
            "year"     : 2026,
            "socio"    : str(socio).strip(),
            "cedula"   : str(cedula).strip() if cedula is not None else "-",
            "cantidad" : safe_float(cantidad),
            "tipo"     : str(tipo).strip() if tipo else "COP",
            "tasa"     : safe_float(tasa_raw)   if isinstance(tasa_raw,   (int, float)) else 0.0,
            "cambio"   : safe_float(cambio_raw) if isinstance(cambio_raw, (int, float)) else 0.0,
        }
        nuevos.append(entry)
        print(f"  + recibo {recibo_str:>4}  {fmt_fecha(fecha_raw)}  {str(socio).strip()[:35]}")

    if not nuevos:
        print("  (sin nuevos movimientos)")
        return mov_data

    return sorted(nuevos, key=lambda x: int(x["recibo"]), reverse=True) + mov_data

# ── Sync 2: COMPORTAMIENTO → patrimonio mensual ───────────────────────────────

def sync_comportamiento(ws, users_map):
    """
    Columnas COMPORTAMIENTO (0-indexed):
      0 = N,  1 = SOCIOS,  2 = DIC anterior
      Por mes i (0=enero … 11=diciembre):
        3+3i = MOVI   4+3i = PATRIMONIO   5+3i = G-P
    """
    cambios = 0
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        n_raw = row[0]
        if n_raw is None or str(n_raw).strip() in ("N", ""):
            continue
        if str(row[1]).strip().upper() == "SOCIOS INACTIVOS":
            break

        nk = n_key_of(n_raw)
        user_info = users_map.get(nk)
        if not user_info:
            continue

        full_path = os.path.join(REPO_PATH, user_info["dataFile"])
        if not os.path.exists(full_path):
            continue

        user_data = parse_userdata(full_path)
        if "meses" not in user_data:
            continue

        modified = False
        for i, mes in enumerate(MESES):
            movi_idx = 3 + 3 * i
            patr_idx = 4 + 3 * i
            gp_idx   = 5 + 3 * i

            if patr_idx >= len(row):
                break

            patr_val = row[patr_idx]
            movi_val = row[movi_idx] if movi_idx < len(row) else None
            gp_val   = row[gp_idx]   if gp_idx   < len(row) else None

            # Solo actualizar meses con actividad real
            movi_real = movi_val is not None and float(movi_val or 0) != 0
            gp_real   = gp_val   is not None and float(gp_val   or 0) != 0
            if not movi_real and not gp_real:
                continue
            if not patr_val:
                continue

            new_patr = round(float(patr_val), 5)
            cur_patr = float(user_data["meses"].get(mes, {}).get("patrimonio", 0) or 0)

            if abs(new_patr - cur_patr) > 0.001:
                if mes not in user_data["meses"]:
                    user_data["meses"][mes] = {"aporte": 0, "patrimonio": 0}
                user_data["meses"][mes]["patrimonio"] = new_patr
                modified = True

        if modified:
            write_userdata(full_path, user_data)
            cambios += 1
            print(f"  ✓ {user_info['username']:<15}  ({user_info['dataFile']})")

    if cambios == 0:
        print("  (sin cambios en patrimonio)")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 52)
    print("  SYNC CASTLE BLACK — Excel → Dashboard")
    print("=" * 52)

    if not os.path.exists(EXCEL_PATH):
        print(f"\nERROR: Excel no encontrado en:\n  {EXCEL_PATH}")
        return

    mtime = datetime.fromtimestamp(os.path.getmtime(EXCEL_PATH))
    print(f"\nExcel: {mtime.strftime('%d/%m/%Y %H:%M:%S')}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    users_map = build_users_map(parse_users(USERS_JS))

    print("\n[1] MOVIMIENTOS — buscando filas nuevas...")
    mov_data = parse_movimientos(MOVIMIENTOS_JS)
    mov_data = sync_movimientos(wb["MOVIMIENTOS"], mov_data, users_map)
    write_movimientos(MOVIMIENTOS_JS, mov_data)
    print(f"  → {len(mov_data)} entradas totales")

    print("\n[2] COMPORTAMIENTO — actualizando patrimonio mensual...")
    sync_comportamiento(wb["COMPORTAMIENTO"], users_map)

    print("\n" + "=" * 52)
    print("  SYNC COMPLETO")
    print("=" * 52 + "\n")

if __name__ == "__main__":
    main()
